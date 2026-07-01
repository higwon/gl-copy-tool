import copy
import os
import re
import sys
import threading
import time
import tkinter as tk
from tkinter import filedialog, messagebox, ttk

from countries import DEFAULT_IMPORT_FORMAT, IMPORT_FORMATS, build_source_records

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
except ImportError:
    DND_FILES = None
    TkinterDnD = None


APP_NAME = "GL Input Copy Tool"
APP_VERSION = "0.4.0"
APP_AUTHOR = "DA_GYEONG"
APP_ICON = "assets/app.ico"
HEADER_LOGO = "assets/logo_header.png"
STATUS_SUCCESS_IMAGE = "assets/status_success.png"
STATUS_ERROR_IMAGE = "assets/status_error.png"
DIALOG_SUCCESS_IMAGE = "assets/dialog_success.png"
DIALOG_ERROR_IMAGE = "assets/dialog_error.png"
TARGET_SHEET_NAME = "2. GL Input"
COA_SHEET_NAME = "1. COA"
ACCOUNT_CODE_HEADER = "계정코드"
MATCH_HEADERS = ["날짜", "계정코드", "차변(EUR)", "대변(EUR)", "거래처명", "적요"]
HEADER_SCAN_ROWS = 30
BaseTk = TkinterDnD.Tk if TkinterDnD else tk.Tk
CELL_REFERENCE_RE = re.compile(r"(?<![A-Za-z0-9_])(\$?[A-Z]{1,3})(\$?)(\d+)")


def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def report_progress(progress_callback, percent, message):
    if progress_callback:
        progress_callback(percent, message)


def is_formula_cell(cell):
    value = cell.value
    return cell.data_type == "f" or (isinstance(value, str) and value.startswith("="))


def is_read_only_merged_cell(cell):
    return cell.__class__.__name__ == "MergedCell"


def normalize_header(value):
    if value is None:
        return ""
    return "".join(str(value).split()).lower()


def find_header_row_and_columns(sheet, required_headers):
    required = {normalize_header(header): header for header in required_headers}

    for row in sheet.iter_rows(max_row=min(sheet.max_row, HEADER_SCAN_ROWS)):
        found = {}
        for cell in row:
            normalized = normalize_header(cell.value)
            if normalized in required and normalized not in found:
                found[normalized] = cell.column

        if all(normalize_header(header) in found for header in required_headers):
            return row[0].row, {
                required[normalized]: column for normalized, column in found.items()
            }

    missing_text = ", ".join(required_headers)
    raise ValueError(
        f'"{sheet.title}" 시트의 상위 {HEADER_SCAN_ROWS}행에서 '
        f"필수 헤더를 찾지 못했습니다: {missing_text}"
    )


def find_formula_columns(sheet):
    formula_columns = set()
    for row in sheet.iter_rows():
        for cell in row:
            if is_read_only_merged_cell(cell):
                continue
            if is_formula_cell(cell):
                formula_columns.add(cell.column)
    return formula_columns


def clear_target_data(sheet, header_row, target_columns, formula_columns, min_clear_row):
    clear_until = max(sheet.max_row, min_clear_row)
    for row in sheet.iter_rows(min_row=header_row + 1, max_row=clear_until):
        for cell in row:
            if is_read_only_merged_cell(cell):
                continue
            if cell.column in target_columns and cell.column not in formula_columns:
                cell.value = None


def build_row_style_template(sheet, source_row_number):
    source_dimension = sheet.row_dimensions[source_row_number]
    row_style = {
        "height": source_dimension.height,
        "hidden": source_dimension.hidden,
        "outlineLevel": source_dimension.outlineLevel,
        "collapsed": source_dimension.collapsed,
    }
    cell_styles = []

    for column_number in range(1, sheet.max_column + 1):
        source_cell = sheet.cell(row=source_row_number, column=column_number)
        cell_styles.append(
            (
                column_number,
                copy.copy(source_cell._style) if source_cell.has_style else None,
            )
        )

    return row_style, cell_styles


def apply_row_style_template(sheet, target_row_number, row_style, cell_styles):
    target_dimension = sheet.row_dimensions[target_row_number]
    target_dimension.height = row_style["height"]
    target_dimension.hidden = row_style["hidden"]
    target_dimension.outlineLevel = row_style["outlineLevel"]
    target_dimension.collapsed = row_style["collapsed"]

    for column_number, style in cell_styles:
        target_cell = sheet.cell(row=target_row_number, column=column_number)
        if is_read_only_merged_cell(target_cell):
            continue
        if target_cell._style != style:
            target_cell._style = style


def expand_target_rows(sheet, header_row, data_row_count, progress_callback=None):
    if data_row_count <= 0:
        return

    first_data_row = header_row + 1
    last_data_row = header_row + data_row_count
    style_source_row = first_data_row
    row_style, cell_styles = build_row_style_template(sheet, style_source_row)
    rows_to_prepare = data_row_count
    for row_number in range(first_data_row, last_data_row + 1):
        apply_row_style_template(sheet, row_number, row_style, cell_styles)
        processed = row_number - header_row
        if processed % 5000 == 0 or processed == rows_to_prepare:
            percent = 42 + int((processed / rows_to_prepare) * 2)
            report_progress(
                progress_callback,
                min(percent, 44),
                f"GL Input 행 서식을 준비하는 중... ({processed}/{rows_to_prepare})",
            )


def adjust_table_ranges(sheet, header_row, data_row_count):
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import range_boundaries

    last_keep_row = header_row + data_row_count
    adjusted_table_count = 0

    for table in sheet.tables.values():
        min_col, min_row, max_col, max_row = range_boundaries(table.ref)
        if not (min_row <= header_row <= max_row or min_row <= last_keep_row <= max_row):
            continue

        new_max_row = max(last_keep_row, min_row)
        new_ref = (
            f"{get_column_letter(min_col)}{min_row}:"
            f"{get_column_letter(max_col)}{new_max_row}"
        )
        if table.ref != new_ref:
            table.ref = new_ref
            if table.autoFilter:
                table.autoFilter.ref = new_ref
            adjusted_table_count += 1

    return adjusted_table_count


def delete_rows_below_input(sheet, header_row, data_row_count):
    first_delete_row = header_row + data_row_count + 1
    last_keep_row = first_delete_row - 1
    deleted_row_count = 0
    if sheet.max_row >= first_delete_row:
        deleted_row_count = sheet.max_row - first_delete_row + 1
        sheet.delete_rows(first_delete_row, deleted_row_count)

    for row_number in list(sheet.row_dimensions):
        if row_number > last_keep_row:
            del sheet.row_dimensions[row_number]

    return deleted_row_count


def translate_formula_row_delta(formula, row_delta):
    if row_delta == 0:
        return formula

    def replace_cell_reference(match):
        column, row_absolute, row_number = match.groups()
        if row_absolute:
            return match.group(0)
        return f"{column}{int(row_number) + row_delta}"

    return CELL_REFERENCE_RE.sub(replace_cell_reference, formula)


def fill_formula_columns(sheet, header_row, data_row_count, formula_columns):
    if data_row_count <= 0:
        return

    first_data_row = header_row + 1
    last_data_row = header_row + data_row_count

    for column in sorted(formula_columns):
        template_cell = None
        for row_number in range(first_data_row, sheet.max_row + 1):
            candidate = sheet.cell(row=row_number, column=column)
            if is_read_only_merged_cell(candidate):
                continue
            if is_formula_cell(candidate):
                template_cell = candidate
                break

        if not template_cell:
            continue

        for row_number in range(first_data_row, last_data_row + 1):
            target_cell = sheet.cell(row=row_number, column=column)
            if is_read_only_merged_cell(target_cell):
                continue

            target_cell.value = translate_formula_row_delta(
                template_cell.value,
                row_number - template_cell.row,
            )

            if template_cell.has_style:
                target_cell._style = copy.copy(template_cell._style)
            if template_cell.number_format:
                target_cell.number_format = template_cell.number_format


def is_empty_value(value):
    return value is None or value == "" or (isinstance(value, str) and not value.strip())


def normalize_account_code_value(value):
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_account_code_number(value):
    if value is None:
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(value) if value.is_integer() else value

    text = str(value).strip()
    if not text:
        return None

    try:
        number = float(text.replace(",", ""))
    except ValueError:
        return text

    return int(number) if number.is_integer() else number


def infer_account_code_mode(template_wb):
    if COA_SHEET_NAME not in template_wb.sheetnames:
        return "text"

    coa_ws = template_wb[COA_SHEET_NAME]
    try:
        header_row, columns = find_header_row_and_columns(coa_ws, [ACCOUNT_CODE_HEADER])
    except ValueError:
        return "text"

    account_code_column = columns[ACCOUNT_CODE_HEADER]
    has_account_code = False
    for row_number in range(header_row + 1, coa_ws.max_row + 1):
        cell = coa_ws.cell(row=row_number, column=account_code_column)
        if is_empty_value(cell.value):
            continue

        has_account_code = True
        if isinstance(cell.value, str) or cell.data_type == "s":
            return "text"

    return "number" if has_account_code else "text"


def coerce_account_code_value(value, account_code_mode):
    if account_code_mode == "number":
        return normalize_account_code_number(value)
    return normalize_account_code_value(value)


def set_target_cell_value(header, value, target_cell, account_code_mode="text"):
    if header == ACCOUNT_CODE_HEADER:
        target_cell.value = coerce_account_code_value(value, account_code_mode)
        if account_code_mode == "number":
            target_cell.number_format = "General"
        else:
            target_cell.number_format = "@"
    else:
        target_cell.value = value


def copy_export_to_gl_input(
    template_path,
    export_path,
    output_path,
    import_format=DEFAULT_IMPORT_FORMAT,
    progress_callback=None,
):
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError as exc:
        raise ValueError(
            "openpyxl이 설치되어 있지 않습니다. PowerShell에서 "
            "'py -m pip install -r requirements.txt'를 실행한 뒤 다시 시도하세요."
        ) from exc

    report_progress(progress_callback, None, "파일을 여는 중...")
    template_wb = load_workbook(template_path)
    export_wb = load_workbook(export_path, data_only=False)

    report_progress(progress_callback, 15, "GL Auto 템플릿 시트를 확인하는 중...")
    if TARGET_SHEET_NAME not in template_wb.sheetnames:
        raise ValueError(f'GL Auto 템플릿에 "{TARGET_SHEET_NAME}" 시트가 없습니다.')

    target_ws = template_wb[TARGET_SHEET_NAME]
    source_ws = export_wb.worksheets[0]

    report_progress(progress_callback, 25, "헤더를 찾는 중...")
    formula_columns = find_formula_columns(target_ws)
    target_header_row, target_columns = find_header_row_and_columns(target_ws, MATCH_HEADERS)
    account_code_mode = infer_account_code_mode(template_wb)

    report_progress(progress_callback, 35, "Source GL을 읽는 중...")
    try:
        source_records = build_source_records(source_ws, import_format, progress_callback)
    finally:
        export_wb.close()

    report_progress(progress_callback, 42, "GL Input 표 범위와 서식을 준비하는 중...")
    expand_target_rows(target_ws, target_header_row, len(source_records), progress_callback)
    adjusted_table_count = adjust_table_ranges(
        target_ws,
        target_header_row,
        len(source_records),
    )

    report_progress(progress_callback, 45, "기존 GL Input 데이터를 삭제하는 중...")
    clear_target_data(
        target_ws,
        target_header_row,
        set(target_columns.values()),
        formula_columns,
        target_header_row + len(source_records),
    )

    total_rows = max(len(source_records), 1)
    for row_offset, source_record in enumerate(source_records, start=1):
        target_row_number = target_header_row + row_offset
        for header in MATCH_HEADERS:
            target_column = target_columns[header]
            if target_column in formula_columns:
                continue

            target_cell = target_ws.cell(row=target_row_number, column=target_column)
            if is_read_only_merged_cell(target_cell):
                continue
            if is_formula_cell(target_cell):
                continue

            set_target_cell_value(
                header,
                source_record.get(header),
                target_cell,
                account_code_mode,
            )

        if row_offset % 250 == 0 or row_offset == len(source_records):
            percent = 45 + int((row_offset / total_rows) * 35)
            report_progress(
                progress_callback,
                percent,
                f"GL Input에 데이터 입력 중... ({row_offset}/{len(source_records)})",
            )

    report_progress(progress_callback, 82, "수식 컬럼을 정리하는 중...")
    fill_formula_columns(target_ws, target_header_row, len(source_records), formula_columns)

    report_progress(progress_callback, 88, "입력 데이터 아래 행을 삭제하는 중...")
    deleted_row_count = delete_rows_below_input(
        target_ws,
        target_header_row,
        len(source_records),
    )

    report_progress(progress_callback, None, "결과 파일을 저장하는 중...")
    template_wb.save(output_path)
    template_wb.close()
    report_progress(progress_callback, 100, "완료")
    return {
        "copied_rows": len(source_records),
        "deleted_rows": deleted_row_count,
        "adjusted_tables": adjusted_table_count,
        "import_format": import_format,
        "output_path": output_path,
    }


class GlInputCopyApp(BaseTk):
    def __init__(self):
        super().__init__()

        self.title(APP_NAME)
        self.geometry("940x720")
        self.minsize(820, 680)
        self.resizable(True, True)
        self.configure(bg="#F3F5F7")

        self._set_window_icon()
        self.header_logo_image = self._load_photo_image(HEADER_LOGO)
        self.status_success_image = self._load_photo_image(STATUS_SUCCESS_IMAGE)
        self.status_error_image = self._load_photo_image(STATUS_ERROR_IMAGE)
        self.dialog_success_image = self._load_photo_image(DIALOG_SUCCESS_IMAGE)
        self.dialog_error_image = self._load_photo_image(DIALOG_ERROR_IMAGE)
        self.status_image_label = None
        self.import_format = tk.StringVar(value=DEFAULT_IMPORT_FORMAT)
        self.template_path = tk.StringVar()
        self.export_path = tk.StringVar()
        self.output_path = tk.StringVar()
        self.status_text = tk.StringVar(value="대기 중")
        self.progress_value = tk.IntVar(value=0)
        self.last_output_path = None
        self.is_running = False
        self.run_started_at = None
        self.current_status_message = "대기 중"
        self.progress_is_indeterminate = False

        self.run_button = None
        self.open_file_button = None
        self.open_folder_button = None
        self.summary_view = None
        self.style = ttk.Style(self)
        self._configure_style()
        self._build_ui()

    def _set_window_icon(self):
        icon_path = resource_path(APP_ICON)
        if os.path.exists(icon_path):
            try:
                self.iconbitmap(icon_path)
            except tk.TclError:
                pass

    def _load_photo_image(self, relative_path):
        path = resource_path(relative_path)
        if not os.path.exists(path):
            return None
        try:
            return tk.PhotoImage(file=path)
        except tk.TclError:
            return None

    def _configure_style(self):
        self.style.theme_use("clam")
        self.style.configure("App.TFrame", background="#F3F5F7")
        self.style.configure("Card.TFrame", background="#FFFFFF", relief="flat")
        self.style.configure(
            "Title.TLabel",
            background="#F3F5F7",
            foreground="#16343B",
            font=("Segoe UI", 18, "bold"),
        )
        self.style.configure(
            "Subtitle.TLabel",
            background="#F3F5F7",
            foreground="#64727A",
            font=("Segoe UI", 9),
        )
        self.style.configure(
            "Meta.TLabel",
            background="#F3F5F7",
            foreground="#64727A",
            font=("Segoe UI", 9),
        )
        self.style.configure(
            "Section.TLabel",
            background="#FFFFFF",
            foreground="#16343B",
            font=("Segoe UI", 10, "bold"),
        )
        self.style.configure(
            "Field.TLabel",
            background="#FFFFFF",
            foreground="#34454C",
            font=("Segoe UI", 9, "bold"),
        )
        self.style.configure(
            "Hint.TLabel",
            background="#FFFFFF",
            foreground="#6B7280",
            font=("Segoe UI", 8),
        )
        self.style.configure(
            "Status.TLabel",
            background="#FFFFFF",
            foreground="#52636B",
            font=("Segoe UI", 9),
        )
        self.style.configure(
            "Summary.TLabel",
            background="#F8FAFC",
            foreground="#374151",
            font=("Segoe UI", 9),
            padding=(10, 8),
        )
        self.style.configure(
            "Primary.TButton",
            font=("Segoe UI", 10, "bold"),
            padding=(18, 10),
            background="#0F766E",
            foreground="#FFFFFF",
        )
        self.style.map(
            "Primary.TButton",
            background=[("active", "#0B5F59"), ("disabled", "#9AA8AC")],
            foreground=[("disabled", "#F3F4F6")],
        )
        self.style.configure(
            "Browse.TButton",
            font=("Segoe UI", 9),
            padding=(11, 7),
            background="#E8EEF0",
            foreground="#294047",
        )
        self.style.map("Browse.TButton", background=[("active", "#DCE6E8")])
        self.style.configure("TEntry", padding=7, fieldbackground="#FBFCFC")
        self.style.configure("TCombobox", padding=6, fieldbackground="#FBFCFC")
        self.style.configure(
            "Blue.Horizontal.TProgressbar",
            troughcolor="#E2E8EA",
            background="#0F766E",
            bordercolor="#E2E8EA",
            lightcolor="#0F766E",
            darkcolor="#0F766E",
        )

    def _build_ui(self):
        container = ttk.Frame(self, padding=(26, 22), style="App.TFrame")
        container.pack(fill=tk.BOTH, expand=True)

        header_frame = ttk.Frame(container, style="App.TFrame")
        header_frame.grid(row=0, column=0, sticky="w")
        if self.header_logo_image:
            ttk.Label(
                header_frame,
                image=self.header_logo_image,
                style="Subtitle.TLabel",
            ).grid(row=0, column=0, rowspan=2, sticky="w", padx=(0, 12))
        ttk.Label(header_frame, text=APP_NAME, style="Title.TLabel").grid(
            row=0, column=1, sticky="w"
        )
        ttk.Label(
            header_frame,
            text="국가별 GL 데이터를 GL Auto 템플릿 형식으로 변환합니다.",
            style="Subtitle.TLabel",
        ).grid(row=1, column=1, sticky="w", pady=(2, 0))
        ttk.Label(
            container,
            text=f"v{APP_VERSION} | {APP_AUTHOR}",
            style="Meta.TLabel",
        ).grid(row=0, column=1, sticky="ne")

        card = ttk.Frame(container, padding=(22, 18), style="Card.TFrame")
        card.grid(row=1, column=0, columnspan=2, sticky="nsew", pady=(18, 0))

        ttk.Label(card, text="입력 설정", style="Section.TLabel").grid(
            row=0, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        ttk.Label(card, text="국가 선택", anchor="w", width=18, style="Field.TLabel").grid(
            row=1, column=0, sticky="w", pady=7
        )
        format_combo = ttk.Combobox(
            card,
            textvariable=self.import_format,
            values=list(IMPORT_FORMATS),
            state="readonly",
            width=34,
        )
        format_combo.grid(row=1, column=1, sticky="w", padx=(8, 8), pady=7)

        self._add_file_row(
            card,
            row=2,
            label="GL Auto 템플릿 xlsx",
            variable=self.template_path,
            command=self.select_template,
            drop_role="template",
        )
        self._add_file_row(
            card,
            row=3,
            label="Source GL xlsx",
            variable=self.export_path,
            command=self.select_export,
            drop_role="export",
        )
        self._add_file_row(
            card,
            row=4,
            label="결과 저장 경로",
            variable=self.output_path,
            command=self.select_output,
            drop_role="output",
        )

        ttk.Separator(card, orient="horizontal").grid(
            row=5, column=0, columnspan=3, sticky="ew", pady=(15, 14)
        )
        ttk.Label(card, text="진행 상태", style="Section.TLabel").grid(
            row=6, column=0, columnspan=3, sticky="w", pady=(0, 8)
        )

        self.progress_bar = ttk.Progressbar(
            card,
            variable=self.progress_value,
            maximum=100,
            mode="determinate",
            style="Blue.Horizontal.TProgressbar",
        )
        self.progress_bar.grid(row=7, column=0, columnspan=3, sticky="ew", pady=(0, 9))

        ttk.Label(card, textvariable=self.status_text, anchor="w", style="Status.TLabel").grid(
            row=8, column=0, columnspan=2, sticky="ew"
        )

        self.run_button = ttk.Button(
            card,
            text="실행",
            command=self.run,
            width=14,
            style="Primary.TButton",
        )
        self.run_button.grid(row=8, column=2, sticky="e")

        ttk.Separator(card, orient="horizontal").grid(
            row=9, column=0, columnspan=3, sticky="ew", pady=(16, 14)
        )
        ttk.Label(card, text="실행 결과", style="Section.TLabel").grid(
            row=10, column=0, columnspan=3, sticky="w", pady=(0, 9)
        )

        summary_frame = ttk.Frame(card, style="Card.TFrame")
        summary_frame.grid(row=11, column=0, columnspan=3, sticky="nsew")

        result_frame = ttk.Frame(summary_frame, style="Card.TFrame")
        result_frame.grid(row=0, column=0, sticky="ew")
        status_frame = tk.Frame(
            result_frame,
            width=66,
            height=100,
            bg="#FFFFFF",
            highlightbackground="#E5E7EB",
            highlightthickness=1,
        )
        status_frame.grid(row=0, column=0, sticky="nsw", padx=(0, 10))
        status_frame.grid_propagate(False)
        self.status_image_label = tk.Label(status_frame, bg="#FFFFFF")
        self.status_image_label.place(relx=0.5, rely=0.5, anchor="center")
        self.summary_view = tk.Text(
            result_frame,
            height=6,
            wrap="word",
            relief="flat",
            bg="#F6F9F9",
            fg="#34454C",
            font=("Segoe UI", 9),
            padx=10,
            pady=8,
        )
        self.summary_view.grid(row=0, column=1, sticky="ew")
        summary_scrollbar = ttk.Scrollbar(
            result_frame,
            orient="vertical",
            command=self.summary_view.yview,
        )
        summary_scrollbar.grid(row=0, column=2, sticky="ns")
        self.summary_view.configure(yscrollcommand=summary_scrollbar.set, state=tk.DISABLED)
        result_frame.grid_columnconfigure(1, weight=1)
        self._set_result_status_image(None)
        self._set_summary_text("실행 결과가 여기에 표시됩니다.")

        button_frame = ttk.Frame(summary_frame, style="Card.TFrame")
        button_frame.grid(row=0, column=1, sticky="e", padx=(12, 0))

        self.open_folder_button = ttk.Button(
            button_frame,
            text="파일 위치 열기",
            command=self.open_output_folder,
            state=tk.DISABLED,
            style="Browse.TButton",
        )
        self.open_folder_button.grid(row=0, column=0, padx=(0, 8))

        self.open_file_button = ttk.Button(
            button_frame,
            text="파일 열기",
            command=self.open_output_file,
            state=tk.DISABLED,
            style="Browse.TButton",
        )
        self.open_file_button.grid(row=0, column=1)

        container.grid_columnconfigure(0, weight=1)
        container.grid_columnconfigure(1, weight=0)
        container.grid_rowconfigure(1, weight=1)
        card.grid_columnconfigure(1, weight=1)
        card.grid_rowconfigure(11, weight=1)
        summary_frame.grid_columnconfigure(0, weight=1)

    def _add_file_row(self, parent, row, label, variable, command, drop_role):
        ttk.Label(parent, text=label, anchor="w", width=18, style="Field.TLabel").grid(
            row=row, column=0, sticky="w", pady=8
        )
        entry = ttk.Entry(parent, textvariable=variable, width=68)
        entry.grid(
            row=row, column=1, sticky="ew", padx=(8, 8), pady=8
        )
        self._register_drop_target(entry, variable, drop_role)
        ttk.Button(parent, text="찾기", command=command, width=10, style="Browse.TButton").grid(
            row=row, column=2, sticky="e", pady=8
        )

    def _register_drop_target(self, widget, variable, drop_role):
        if not DND_FILES:
            return

        try:
            widget.drop_target_register(DND_FILES)
            widget.dnd_bind(
                "<<Drop>>",
                lambda event: self._handle_file_drop(event, variable, drop_role),
            )
        except tk.TclError:
            return

    def _handle_file_drop(self, event, variable, drop_role):
        paths = [path for path in self.tk.splitlist(event.data) if path]
        if not paths:
            return

        path = paths[0]
        if drop_role == "output" and os.path.isdir(path):
            base_name = "GL_Input_Result"
            if self.template_path.get().strip():
                base_name = os.path.splitext(os.path.basename(self.template_path.get().strip()))[0]
            variable.set(os.path.join(path, f"{base_name}_result.xlsx"))
            return

        if not path.lower().endswith(".xlsx"):
            messagebox.showerror("오류", "xlsx 파일만 드래그앤드랍할 수 있습니다.")
            return

        variable.set(path)
        if drop_role in ("template", "export"):
            self._suggest_output_path()

    def _set_summary_text(self, text):
        if not self.summary_view:
            return
        self.summary_view.configure(state=tk.NORMAL)
        self.summary_view.delete("1.0", tk.END)
        self.summary_view.insert("1.0", text)
        self.summary_view.configure(state=tk.DISABLED)

    def _set_result_status_image(self, image):
        if not self.status_image_label:
            return
        if image:
            self.status_image_label.configure(image=image, text="")
        else:
            self.status_image_label.configure(image="", text="Ready", fg="#6B7280")

    def _show_image_dialog(self, title, message, image):
        dialog = tk.Toplevel(self)
        dialog.title(title)
        dialog.configure(bg="#FFFFFF")
        dialog.resizable(False, False)
        dialog.transient(self)
        dialog.grab_set()

        body = ttk.Frame(dialog, padding=20, style="Card.TFrame")
        body.grid(row=0, column=0, sticky="nsew")
        if image:
            ttk.Label(body, image=image, style="Field.TLabel").grid(
                row=0, column=0, sticky="n", padx=(0, 18)
            )

        message_frame = ttk.Frame(body, style="Card.TFrame")
        message_frame.grid(row=0, column=1, sticky="nsew")
        ttk.Label(
            message_frame,
            text=title,
            style="Field.TLabel",
        ).grid(row=0, column=0, sticky="w")
        message_view = tk.Text(
            message_frame,
            width=58,
            height=7,
            wrap="word",
            relief="flat",
            bg="#FFFFFF",
            fg="#374151",
            font=("Segoe UI", 9),
        )
        message_view.grid(row=1, column=0, sticky="ew", pady=(8, 12))
        message_view.insert("1.0", message)
        message_view.configure(state=tk.DISABLED)

        ok_button = ttk.Button(
            message_frame,
            text="확인",
            command=dialog.destroy,
            width=12,
            style="Primary.TButton",
        )
        ok_button.grid(row=2, column=0, sticky="e")

        dialog.update_idletasks()
        x = self.winfo_rootx() + (self.winfo_width() - dialog.winfo_width()) // 2
        y = self.winfo_rooty() + (self.winfo_height() - dialog.winfo_height()) // 2
        dialog.geometry(f"+{max(x, 0)}+{max(y, 0)}")
        ok_button.focus_set()
        self.wait_window(dialog)

    def select_template(self):
        path = filedialog.askopenfilename(
            title="GL Auto 템플릿 xlsx 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.template_path.set(path)
            self._suggest_output_path()

    def select_export(self):
        path = filedialog.askopenfilename(
            title="Source GL xlsx 선택",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.export_path.set(path)
            self._suggest_output_path()

    def select_output(self):
        path = filedialog.asksaveasfilename(
            title="결과 저장 경로 선택",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self.output_path.set(path)

    def _suggest_output_path(self):
        if self.output_path.get() or not self.template_path.get():
            return

        base_dir = os.path.dirname(self.template_path.get())
        base_name = os.path.splitext(os.path.basename(self.template_path.get()))[0]
        self.output_path.set(os.path.join(base_dir, f"{base_name}_result.xlsx"))

    def validate_inputs(self):
        import_format = self.import_format.get().strip()
        template_path = self.template_path.get().strip()
        export_path = self.export_path.get().strip()
        output_path = self.output_path.get().strip()

        if import_format not in IMPORT_FORMATS:
            raise ValueError("국가를 선택하세요.")
        if not template_path:
            raise ValueError("GL Auto 템플릿 xlsx를 선택하세요.")
        if not export_path:
            raise ValueError("Source GL xlsx를 선택하세요.")
        if not output_path:
            raise ValueError("결과 저장 경로를 선택하세요.")
        if not os.path.isfile(template_path):
            raise ValueError("GL Auto 템플릿 파일을 찾을 수 없습니다.")
        if not os.path.isfile(export_path):
            raise ValueError("Source GL 파일을 찾을 수 없습니다.")
        normalized_output = os.path.normcase(os.path.abspath(output_path))
        if os.path.normcase(os.path.abspath(template_path)) == normalized_output:
            raise ValueError("결과 파일은 GL Auto 템플릿과 다른 경로로 저장하세요.")
        if os.path.normcase(os.path.abspath(export_path)) == normalized_output:
            raise ValueError("결과 파일은 Source GL과 다른 경로로 저장하세요.")

        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.isdir(output_dir):
            raise ValueError("결과 저장 폴더를 찾을 수 없습니다.")

        return import_format, template_path, export_path, output_path

    def update_progress(self, percent, message):
        self.after(0, self._set_progress, percent, message)

    def _set_progress(self, percent, message):
        self.current_status_message = message
        if percent is None:
            if not self.progress_is_indeterminate:
                self.progress_bar.configure(mode="indeterminate")
                self.progress_bar.start(14)
                self.progress_is_indeterminate = True
        else:
            if self.progress_is_indeterminate:
                self.progress_bar.stop()
                self.progress_bar.configure(mode="determinate")
                self.progress_is_indeterminate = False
            self.progress_value.set(percent)
        self._refresh_running_status()

    def _refresh_running_status(self):
        if not self.is_running or not self.run_started_at:
            self.status_text.set(self.current_status_message)
            return

        elapsed_seconds = int(time.monotonic() - self.run_started_at)
        minutes, seconds = divmod(elapsed_seconds, 60)
        self.status_text.set(
            f"{self.current_status_message} (경과 {minutes:02d}:{seconds:02d})"
        )

    def _tick_running_status(self):
        if not self.is_running:
            return
        self._refresh_running_status()
        self.after(1000, self._tick_running_status)

    def set_running(self, is_running):
        state = tk.DISABLED if is_running else tk.NORMAL
        if self.run_button:
            self.run_button.config(state=state)
        if is_running:
            self._set_result_buttons_enabled(False)

    def run(self):
        try:
            import_format, template_path, export_path, output_path = self.validate_inputs()
        except Exception as exc:
            self.progress_value.set(0)
            self.status_text.set("실패")
            self._set_summary_text(f"실패: 실행 전 확인이 필요합니다.\n원인: {exc}")
            self._set_result_status_image(self.status_error_image)
            self._set_result_buttons_enabled(False)
            self._show_image_dialog("오류", str(exc), self.dialog_error_image)
            return

        self.progress_value.set(0)
        self.status_text.set("시작하는 중...")
        self.current_status_message = "시작하는 중..."
        self._set_summary_text("실행 중입니다. 잠시만 기다려 주세요.")
        self._set_result_status_image(None)
        self.last_output_path = None
        self._set_result_buttons_enabled(False)
        self.is_running = True
        self.run_started_at = time.monotonic()
        self.set_running(True)
        self._tick_running_status()

        worker = threading.Thread(
            target=self._run_worker,
            args=(import_format, template_path, export_path, output_path),
            daemon=True,
        )
        worker.start()

    def _run_worker(self, import_format, template_path, export_path, output_path):
        try:
            result = copy_export_to_gl_input(
                template_path,
                export_path,
                output_path,
                import_format=import_format,
                progress_callback=self.update_progress,
            )
            self.after(0, self._show_success, result)
        except Exception as exc:
            self.after(0, self._show_error, str(exc))

    def _show_success(self, result):
        output_path = result["output_path"]
        self.is_running = False
        self.run_started_at = None
        if self.progress_is_indeterminate:
            self.progress_bar.stop()
            self.progress_bar.configure(mode="determinate")
            self.progress_is_indeterminate = False
        self.set_running(False)
        self.progress_value.set(100)
        self.status_text.set("완료")
        self.current_status_message = "완료"
        self.last_output_path = output_path
        self._set_result_status_image(self.status_success_image)
        self._set_summary_text(
            "성공: GL Input 데이터 입력을 완료했습니다.\n"
            f"국가 선택: {result['import_format']}\n"
            f"입력 행 수: {result['copied_rows']} / "
            f"삭제된 하단 행 수: {result['deleted_rows']} / "
            f"조정된 표 범위: {result['adjusted_tables']}\n"
            f"결과 파일: {output_path}"
        )
        self._set_result_buttons_enabled(True)
        self._show_image_dialog(
            "완료",
            f"결과 파일을 저장했습니다.\n\n{output_path}",
            self.dialog_success_image,
        )

    def _show_error(self, message):
        self.is_running = False
        self.run_started_at = None
        if self.progress_is_indeterminate:
            self.progress_bar.stop()
            self.progress_bar.configure(mode="determinate")
            self.progress_is_indeterminate = False
        self.set_running(False)
        self.progress_value.set(0)
        self.status_text.set("실패")
        self.current_status_message = "실패"
        self.last_output_path = None
        self._set_result_status_image(self.status_error_image)
        self._set_summary_text(f"실패: 작업을 완료하지 못했습니다.\n원인: {message}")
        self._set_result_buttons_enabled(False)
        self._show_image_dialog("오류", message, self.dialog_error_image)

    def _set_result_buttons_enabled(self, enabled):
        state = tk.NORMAL if enabled else tk.DISABLED
        if self.open_file_button:
            self.open_file_button.config(state=state)
        if self.open_folder_button:
            self.open_folder_button.config(state=state)

    def open_output_file(self):
        if self.last_output_path and os.path.exists(self.last_output_path):
            os.startfile(self.last_output_path)
        else:
            messagebox.showerror("오류", "열 수 있는 결과 파일이 없습니다.")

    def open_output_folder(self):
        if self.last_output_path:
            folder_path = os.path.dirname(self.last_output_path)
            if os.path.isdir(folder_path):
                os.startfile(folder_path)
                return
        messagebox.showerror("오류", "열 수 있는 결과 파일 위치가 없습니다.")


if __name__ == "__main__":
    app = GlInputCopyApp()
    app.mainloop()
